import re
from openpyxl import load_workbook
import openpyxl

class digest:
    def __init__(self, document):
        self.document = document
    def grepData(self, str):
        ## search for the date format of xx/xx/xxxx and greps it
        result = re.search("([0-9]{2}\/[0-9]{2}\/[0-9]{4})", str)
        return result.group(1)
    def grepHour(self, str):
        ## search for the hour format of xx:xx and greps it
        result = re.search("([0-9]{2}\:[0-9]{2})", str)
        return result.group(1)
    def populateData(self):
        wbXL = load_workbook(r""+self.document, data_only=True)
        sh = wbXL[wbXL.sheetnames[0]]
        ws = wbXL.active
        rows = sh.max_row
        cols = sh.max_column
        dataMap = []
        placeholder = []
        hourlist = []
        for i in range(rows+1):
            if i >= 3:
                ##generate the datamap (array of array of incidents per date/hour)
                _str = str(ws.cell(row=i, column=3).value)
                date = self.grepData(_str)
                hour = self.grepHour(_str)
                date = date[:5]
                hour = hour[:2]
                if not hour in hourlist:
                    hourlist.append(hour)
                if not date in placeholder:
                    placeholder.append(date)
                    dataMap.append([date, hour])
                else:
                    j = placeholder.index(date)
                    dataMap[j].append(hour)
                ## sort dates and hours of each array inside the datamap
        for i in range(0,len(dataMap)):
            dataMap[i] = dataMap[i][::-1] #reverses the array
            temp = dataMap[i][-1] #grabs the last item (date) from the array
            del(dataMap[i][-1]) #deletes the last item (date) of the array
            dataMap[i].insert(0, temp) #insert the previous saved date to the first position of the array
        dataMap = dataMap[::-1] # reverses the order of the arrays (for some reason CA generates it backwards last -> first)
        self.hourlist = sorted(hourlist)
        self.dataMap = dataMap
    def generateReport(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ## populate the columns of hours
        initrow = 3
        initcol = 2
        for i in self.hourlist:
            ws.cell(row=initrow, column=initcol, value=i)
            initrow+= 1
        ## populate the rows of date
        initrow= 2
        initcol += 1
        for j in self.dataMap:
            ws.cell(row=initrow, column=initcol, value=j[0])
            initcol += 1
        ## populate the contents of the spreadsheet
        initcol = 3
        initrow = 2
        for i in self.dataMap:
            for j in i:
                if j in self.hourlist:
                    indexHour = self.hourlist.index(j)+1
                    val = ws.cell(row=initrow+indexHour, column=initcol).value
                    if val is not None:
                        val = int(val)+1
                    else:
                        val = 1
                    ws.cell(row=initrow+indexHour, column=initcol, value=val)
            initcol += 1


        wb.save("sample.xlsx")